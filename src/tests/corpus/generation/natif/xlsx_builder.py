# ====== Code Summary ======
# XlsxCorpusBuilder — builds a LONG, complex, multilingual .xlsx from the report content pack:
# a Data sheet with many styled rows (merged title, header fills, number formats, frozen panes,
# a totals row of formulas), a Summary sheet of cross-sheet formulas/KPIs, a Commentary sheet of
# real multilingual prose (gives language-detection signal + stresses chunking once rendered), and
# a Visual sheet with a native chart + an embedded image. The volume of rows yields multiple PDF
# pages after LibreOffice conversion. Driven by spec.language (content pulled from the report pack).

# ====== Standard Library Imports ======
from __future__ import annotations

import datetime
import io

# ====== Third-Party Library Imports ======
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ====== Local Project Imports ======
from .base import BaseDocumentBuilder
from .content import get_content
from .image_factory import ImageFactory

# Number of synthetic monthly periods per region — drives row count (and rendered page count).
_PERIODS = 12
# Localized sheet names per language.
_SHEETS = {
    "fr": ("Données", "Synthèse", "Commentaire", "Visuel"),
    "en": ("Data", "Summary", "Commentary", "Visual"),
    "es": ("Datos", "Resumen", "Comentario", "Visual"),
}
_HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
_TITLE_FONT = Font(bold=True, size=14)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


class XlsxCorpusBuilder(BaseDocumentBuilder):
    """Builds a long, complex, multilingual spreadsheet from the report content pack."""

    def build(self) -> bytes:
        """
        Assemble the .xlsx and return its bytes.

        Returns:
            bytes: A long, multi-sheet .xlsx workbook.
        """
        # 1. Resolve multilingual content (report pack) + sheet names
        content = get_content("report", self.spec.language or "fr")
        names = _SHEETS.get(self.spec.language, _SHEETS["en"])
        wb = Workbook()

        # 2. Build each sheet
        data_ws = wb.active
        data_ws.title = names[0]
        last_row = self.__build_data_sheet(data_ws, content)
        self.__build_summary_sheet(wb.create_sheet(names[1]), content, names[0], last_row)
        self.__build_commentary_sheet(wb.create_sheet(names[2]), content)
        self.__build_visual_sheet(wb.create_sheet(names[3]), data_ws, names[0], last_row)

        # 3. Serialize
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def __build_data_sheet(self, ws: object, content: object) -> int:
        """Write a merged title, a styled header, many synthetic rows + a totals formula row."""
        # 1. Merged title carrying the searchable phrase (language signal)
        headers = ["Période", *content.table_headers]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value=f"{content.title} — {content.searchable_phrase}").font = _TITLE_FONT
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # 2. Styled header row
        for col, head in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=head)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.border = _BORDER

        # 3. Many data rows: each catalog region replicated across monthly periods with varied numbers
        row = 3
        for period in range(1, _PERIODS + 1):
            label = datetime.date(2026, period, 1).strftime("%Y-%m")
            for r_idx, seed in enumerate(content.table_rows):
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=seed[0])  # region label
                ws.cell(row=row, column=3, value=1000 + period * 37 + r_idx * 11).number_format = "#,##0"
                ws.cell(row=row, column=4, value=400 + period * 13 + r_idx * 7).number_format = "#,##0"
                ws.cell(row=row, column=5, value=round((period + r_idx) / 100, 3)).number_format = "0.0%"
                ws.cell(row=row, column=6, value=20 + r_idx).number_format = "0"
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).border = _BORDER
                row += 1

        # 4. Totals row with SUM formulas; freeze the header
        ws.cell(row=row, column=1, value="Σ").font = Font(bold=True)
        for col in (3, 4, 6):
            letter = ws.cell(row=row, column=col).column_letter
            ws.cell(row=row, column=col, value=f"=SUM({letter}3:{letter}{row - 1})").font = Font(bold=True)
        ws.freeze_panes = "A3"
        return row

    @staticmethod
    def __build_summary_sheet(ws: object, content: object, data_sheet: str, last_row: int) -> None:
        """Cross-sheet KPI formulas (SUM / AVERAGE / COUNT) referencing the data sheet."""
        ws["A1"] = content.section_title(3)
        ws["A1"].font = _TITLE_FONT
        kpis = [
            ("Total revenu", f"=SUM('{data_sheet}'.C3:C{last_row - 1})"),
            ("Total coût", f"=SUM('{data_sheet}'.D3:D{last_row - 1})"),
            ("Revenu moyen", f"=AVERAGE('{data_sheet}'.C3:C{last_row - 1})"),
            ("Lignes", f"=COUNT('{data_sheet}'.C3:C{last_row - 1})"),
        ]
        for i, (label, formula) in enumerate(kpis, start=3):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=formula).number_format = "#,##0.00"

    @staticmethod
    def __build_commentary_sheet(ws: object, content: object) -> None:
        """Real multilingual prose (abstract + paragraphs) — language signal + chunking material."""
        ws.column_dimensions["A"].width = 120
        ws["A1"] = content.title
        ws["A1"].font = _TITLE_FONT
        ws["A2"] = content.abstract
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        for i in range(8):
            cell = ws.cell(row=3 + i, column=1, value=content.para(i))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    @staticmethod
    def __build_visual_sheet(ws: object, data_ws: object, data_sheet: str, last_row: int) -> None:
        """A native bar + line chart over the data, plus an embedded synthetic image."""
        ws["A1"] = "KPI"
        ws["A1"].font = _TITLE_FONT
        # Chart over the first 12 data rows (revenue + cost columns)
        end = min(last_row - 1, 14)
        data_ref = Reference(data_ws, min_col=3, max_col=4, min_row=2, max_row=end)
        cats = Reference(data_ws, min_col=1, min_row=3, max_row=end)
        bar = BarChart()
        bar.title = "Revenu vs Coût"
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "A3")
        line = LineChart()
        line.title = "Tendance"
        line.add_data(data_ref, titles_from_data=True)
        ws.add_chart(line, "A20")
        ws.add_image(XlImage(io.BytesIO(ImageFactory.diagram())), "K3")
